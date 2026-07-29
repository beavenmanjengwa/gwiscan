#!/usr/bin/env python3
"""
####################################################################################################
#                                                                                                  #
# compile.py - Join all annotation tables into the final TSV + XLSX (the `compile` stage).         #
#                                                                                                  #
# Base table is the merged candidate hits (one row per domain hit); protein-level annotations      #
# (protparam, targetp, deeptmhmm, deeploc) are left-joined on protein_id, and InterProScan is      #
# summarised per protein before joining. Superfamily mode adds a superfamily column + rollup.      #
#                                                                                                  #
####################################################################################################
"""

from __future__ import annotations

import pandas as pd

from . import external, go, io
from .config import Config

# Grouped top-header bands for the XLSX All_candidates sheet: each column is
# placed under a band naming the tool / data type it comes from. Ordered by
# native (snake_case) column name -- matched before io.to_camel is applied, so
# the irregular TargetP class columns (noTP/SP/mTP/cTP/luTP) match as-is. A
# column not listed here falls under a blank band (no crash); the order of BANDS
# only labels contiguous runs, it does not reorder columns. TSV is unaffected --
# a two-row header would break every parser (incl. our own io.read_tsv), so the
# grouping lives only in the human-facing spreadsheet.
COLUMN_BANDS = [
    ("Identity", ["protein_id", "family", "superfamily"]),
    ("Chromosomal Localization", ["gene_id", "chrom", "gene_start", "gene_end", "strand"]),
    ("Search hit (hmmscan / BLAST)", ["accession", "evalue", "bitscore", "start", "end", "method"]),
    ("Domain architecture", ["domain_architecture", "family_domain_count", "architecture_type"]),
    ("Evidence", ["evidence_level", "evidence_support", "evidence_criteria"]),
    ("Computed Physicochemical Properties",
     ["length_aa", "molecular_weight", "isoelectric_point", "negatively_charged_residues",
      "positively_charged_residues", "instability_index", "gravy", "aliphatic_index",
      "ec_cystines", "ec_reduced"]),
    ("Signal / transit peptide",
     ["targetp_type", "noTP", "SP", "mTP", "cTP", "luTP", "cs_position"]),
    ("Transmembrane topology",
     ["topology", "signal_peptide", "tm_regions", "n_tm_regions", "beta_regions"]),
    ("Predicted Subcellular localization", ["localizations", "signals", "membrane_types"]),
    ("Domains & InterPro2GO",
     ["ipr_accessions", "ipr_descriptions", "go_terms", "go_term_names", "go_molecular_function",
      "go_biological_process", "go_cellular_component", "analyses_hit"]),
]

# Soft fills cycled across bands so adjacent groups read as distinct blocks.
_BAND_FILLS = ["DDEBF7", "E2EFDA", "FCE4D6", "FFF2CC", "EAD1DC", "D9E1F2", "F8CBAD"]


def band_spans(columns) -> list:
    """Contiguous (label, first_idx, last_idx) spans over ``columns`` (0-indexed).

    Each column is labelled by the band it belongs to (COLUMN_BANDS), then equal
    adjacent labels are merged into one span. Columns not in any band get a blank
    label and still occupy their own (unmerged) span, so nothing is dropped and
    the spans always tile the full width exactly once -- pure and testable without
    touching a spreadsheet."""
    col_to_band = {col: label for label, cols in COLUMN_BANDS for col in cols}
    labels = [col_to_band.get(str(c), "") for c in columns]

    spans, i = [], 0
    while i < len(labels):
        j = i
        while j + 1 < len(labels) and labels[j + 1] == labels[i]:
            j += 1
        spans.append((labels[i], i, j))
        i = j + 1
    return spans


def _apply_band_header(ws, spans) -> None:
    """Write the merged band row (Excel row 1) above the column-name row (row 2)
    on an openpyxl worksheet, style both, and freeze them. The sheet must have
    been written with startrow=1 so row 1 is free for the bands."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="BFBFBF")
    band_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_idx = 0
    for label, first, last in spans:
        c0, c1 = first + 1, last + 1                    # openpyxl is 1-indexed
        top = ws.cell(row=1, column=c0, value=label)
        if last > first:
            ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        fill = PatternFill("solid", fgColor=_BAND_FILLS[fill_idx % len(_BAND_FILLS)]) if label else None
        fill_idx += 1 if label else 0
        for col in range(c0, c1 + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = band_font
            cell.alignment = center
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if fill is not None:
                cell.fill = fill
        # the actual column-name row (row 2) shares the band's tint, bolded
        for col in range(c0, c1 + 1):
            hdr = ws.cell(row=2, column=col)
            hdr.font = Font(bold=True)
            hdr.alignment = Alignment(horizontal="center")
            hdr.border = Border(bottom=thin)
            if fill is not None:
                hdr.fill = fill

    # Keep both header rows visible; freeze the first two id columns too so the
    # 50+ annotation columns scroll under a fixed proteinId/family label.
    ws.freeze_panes = f"{get_column_letter(min(3, ws.max_column))}3"


def _load(path, label):
    if path.exists():
        df = io.read_tsv(path, low_memory=False)
        external.log(f"[OK] Loaded {label}: {len(df)} rows, columns: {list(df.columns)[:6]}...")
        return df
    external.log(f"[WARN] {label} not found, skipping: {path}")
    return None


def pfam_architecture(interpro_df) -> dict:
    """Per-protein Pfam domain architecture, ordered N->C by start position.

    Only Pfam matches are used, so every domain is in ONE coordinate system -- the
    same Pfam boundaries the pipeline extracts the family domain from. Returns
    {protein_id: [(bare_pfam_acc, label), ...]}; label is the signature description
    (or the accession if none)."""
    arch = {}
    if interpro_df is None or len(interpro_df) == 0:
        return arch
    pfam = interpro_df[interpro_df["analysis"].astype(str) == "Pfam"]
    for pid, group in pfam.groupby("protein_id"):
        rows = []
        for _, r in group.iterrows():
            try:
                start = int(r["start"])
            except (ValueError, TypeError):
                continue
            acc = str(r["sig_acc"]).split(".")[0]
            desc = str(r["sig_desc"])
            rows.append((start, acc, desc if desc not in ("-", "nan", "") else acc))
        arch[str(pid)] = [(acc, label) for _, acc, label in sorted(rows)]
    return arch


def evidence_support(merged_df) -> dict:
    """{(protein_id, family): 'both' | 'hmm_only' | 'blast_only'} from the merged hits.

    Which of the two independent searches recovered each protein for each family:
    the per-protein counterpart of the per-family detectability profile."""
    support = {}
    if merged_df is None or len(merged_df) == 0:
        return support
    for (pid, family), group in merged_df.groupby(["protein_id", "family"]):
        methods = {str(m).lower() for m in group["method"]}
        both = "hmm" in methods and "blast" in methods
        support[(str(pid), str(family))] = (
            "both" if both else "hmm_only" if "hmm" in methods else "blast_only")
    return support


def support_databases(interpro_df) -> dict:
    """{protein_id: number of distinct InterProScan member databases annotating it}."""
    if interpro_df is None or len(interpro_df) == 0:
        return {}
    counts = interpro_df.groupby("protein_id")["analysis"].nunique()
    return {str(pid): int(n) for pid, n in counts.items()}


def confidence(support: str, n_databases: int) -> tuple:
    """(criteria, level) for one candidate from the methods that support the call.

    One criterion per independent source of evidence -- no weights, no tuning:
      * hmmProfile     -- an hmmscan profile hit at the gathering threshold.
      * blastHit       -- a DIAMOND hit.
      * interproDomain -- an InterProScan domain annotation.
    The level carries one '+' per source (+++ / ++ / +), so the mark states how many
    methods independently support the call, and the criteria naming them sit in the
    next column."""
    met = []
    if support in ("both", "hmm_only"):
        met.append("hmmProfile")
    if support in ("both", "blast_only"):
        met.append("blastHit")
    if n_databases >= 1:
        met.append("interproDomain")
    return ("+".join(met) or "-", "+" * len(met) or "-")


def architecture_type(pfam_accs, family_pfam) -> tuple:
    """(family_domain_count, 'standalone' | 'multidomain' | '-') for a protein's
    ordered Pfam accessions, relative to its family's own Pfam accession. Only the
    family domain present (any number of copies) -> standalone; any other domain
    -> multidomain; no Pfam evidence -> '-'."""
    if not pfam_accs:
        return (0, "-")
    fam_count = sum(1 for a in pfam_accs if a == family_pfam) if family_pfam else 0
    has_other = any(a != family_pfam for a in pfam_accs)
    return (fam_count, "multidomain" if has_other else "standalone")


def run(cfg: Config) -> None:
    """Join the confirmed candidates with all annotations into the final outputs."""
    cfg.final_dir.mkdir(parents=True, exist_ok=True)
    final = cfg.result("final_candidates.tsv")   # InterProScan-confirmed candidates
    out_tsv = cfg.final_dir / "gwiscan_results.tsv"
    out_xlsx = cfg.final_dir / "gwiscan_results.xlsx"

    df = io.read_tsv(final, low_memory=False)
    df["protein_id"] = df["protein_id"].astype(str)
    external.log(f"[OK] Base candidates: {len(df)} domain hits, {df['protein_id'].nunique()} proteins")

    # Superfamily mode: map each family to its superfamily for grouping/rollup.
    # Grouping only — no processing differs between modes.
    superfamily_mode = str(cfg.MODE).lower() == "superfamily"
    if superfamily_mode:
        fam2super = io.family_to_superfamily(cfg.family_map)
        if not fam2super:
            external.log("[WARN] MODE=superfamily but the family table has no "
                         "Superfamily column/values; skipping the superfamily rollup.")
            superfamily_mode = False
        else:
            df.insert(df.columns.get_loc("family") + 1, "superfamily",
                      df["family"].map(fam2super).fillna("-"))

    # Chromosomal coordinates identify the locus, so they join the identity block
    # ahead of the domain hit. Absent when no annotation was supplied.
    loc = _load(cfg.result("chromosome_map.tsv"), "chromosome map")
    if loc is not None:
        loc["protein_id"] = loc["protein_id"].astype(str)
        loc = loc.drop(columns=[c for c in ("family",) if c in loc.columns])
        df = df.merge(loc, on="protein_id", how="left")
        # A member the annotation does not resolve leaves an empty coordinate, so
        # the columns are nullable integers rather than floats.
        for column in ("gene_start", "gene_end"):
            if column in df.columns:
                df[column] = pd.to_numeric(df[column], errors="coerce").astype("Int64")
        at = df.columns.get_loc("family") + 1
        for offset, column in enumerate(("gene_id", "chrom", "gene_start",
                                        "gene_end", "strand")):
            if column in df.columns:
                df.insert(at + offset, column, df.pop(column))

    pp = _load(cfg.result("protparam.tsv"), "protparam")
    tp = _load(cfg.result("targetp.tsv"), "targetp")
    tm = _load(cfg.result("deeptmhmm.tsv"), "deeptmhmm")
    dl = _load(cfg.result("deeploc.tsv"), "deeploc")

    for ann in (pp, tp, tm, dl):
        if ann is not None:
            ann["protein_id"] = ann["protein_id"].astype(str)
            df = df.merge(ann, on="protein_id", how="left")

    ipr = _load(cfg.result("interproscan.tsv"), "interproscan")
    if ipr is not None:
        ipr["protein_id"] = ipr["protein_id"].astype(str)

        def _join_unique(values):
            return "|".join(sorted({str(v) for v in values if str(v) not in ("-", "nan", "")}))

        def _join_go(values):
            terms = {gid for v in values for gid in go.clean_ids(v)}
            return "|".join(sorted(terms))

        ipr_summary = ipr.groupby("protein_id").agg(
            ipr_accessions=("ipr_acc", _join_unique),
            ipr_descriptions=("ipr_desc", _join_unique),
            go_terms=("go_terms", _join_go),
            analyses_hit=("analysis", lambda x: "|".join(sorted({str(v) for v in x}))),
        ).reset_index()
        df = df.merge(ipr_summary, on="protein_id", how="left")

    # GO ids -> names + aspect split (InterProScan reports ids only; go-basic.obo
    # supplies the name and namespace for each id).
    if "go_terms" in df.columns:
        go_terms = go.load_terms(cfg)
        if go_terms:
            at = df.columns.get_loc("go_terms") + 1
            df.insert(at, "go_term_names",
                      df["go_terms"].map(lambda v: go.names_for(v, go_terms)))
            aspects = df["go_terms"].map(lambda v: go.names_by_aspect(v, go_terms))
            for offset, aspect in enumerate(go.ASPECTS, start=1):
                df.insert(at + offset, f"go_{aspect}",
                          aspects.map(lambda d, a=aspect: d[a]))

    # Pfam-based domain architecture (N->C) + standalone/multidomain flag. Uses
    # Pfam only (one coordinate system); the family domain is known per row from the
    # family's own Pfam accession.
    arch = pfam_architecture(ipr)
    fam_to_pfam = {r["family"]: r["pfam_model"]
                   for r in io.family_records(cfg.family_map) if r["pfam_model"]}
    architectures = [" + ".join(label for _, label in arch.get(str(pid), [])) or "-"
                     for pid in df["protein_id"]]
    counts, types = [], []
    for pid, family in zip(df["protein_id"], df["family"]):
        accs = [acc for acc, _ in arch.get(str(pid), [])]
        n, kind = architecture_type(accs, fam_to_pfam.get(family))
        counts.append(n)
        types.append(kind)

    # Architecture describes the domains, so it stays with the domain columns
    # rather than trailing the annotation block that is merged in before it.
    at = df.columns.get_loc("method") + 1
    for offset, (name, values) in enumerate(
            [("domain_architecture", architectures),
             ("family_domain_count", counts),
             ("architecture_type", types)]):
        df.insert(at + offset, name, values)

    # Per-candidate evidence: which of hmmscan, DIAMOND and InterProScan support
    # the call, marked one '+' per method.
    merged = _load(cfg.result("candidates_merged.tsv"), "merged candidates")
    support = evidence_support(merged)
    db_counts = support_databases(ipr)
    supports, criteria, levels = [], [], []
    for pid, family in zip(df["protein_id"], df["family"]):
        found_by = support.get((str(pid), str(family)), "-")
        crit, level = confidence(found_by, db_counts.get(str(pid), 0))
        supports.append(found_by)
        criteria.append(crit)
        levels.append(level)
    # The evidence states how the call was made, so it closes the identification
    # block (identity, domain hit, architecture) ahead of the annotation columns.
    at = df.columns.get_loc("architecture_type") + 1
    for offset, (name, values) in enumerate(
            [("evidence_level", levels),
             ("evidence_support", supports),
             ("evidence_criteria", criteria)]):
        df.insert(at + offset, name, values)

    df.sort_values(["family", "protein_id", "evalue"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    io.write_df(df, out_tsv, "tsv")
    external.log(f"[OK] Final TSV saved: gwiscan_results.tsv ({len(df)} rows, {df['protein_id'].nunique()} proteins)")

    # XLSX sheets — camelCase headers (io.to_camel) to match the TSV. The main
    # sheet gets a second, merged header row above the column names, banding each
    # column under the tool / data type it came from (band_spans / _apply_band_header).
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        spans = band_spans(df.columns)
        df.rename(columns=io.to_camel).to_excel(
            writer, sheet_name="All_candidates", index=False, startrow=1)
        _apply_band_header(writer.sheets["All_candidates"], spans)

        summary = df.groupby("family").agg(
            n_proteins=("protein_id", "nunique"),
            n_domain_hits=("protein_id", "count"),
            method=("method", "first"),
            mean_bitscore=("bitscore", "mean"),
            min_evalue=("evalue", "min"),
        ).round(4).reset_index()
        summary.rename(columns=io.to_camel).to_excel(writer, sheet_name="Family_summary", index=False)

        # Superfamily rollup (superfamily mode only): proteins/hits per superfamily.
        if superfamily_mode:
            super_summary = df.groupby("superfamily").agg(
                n_families=("family", "nunique"),
                n_proteins=("protein_id", "nunique"),
                n_domain_hits=("protein_id", "count"),
            ).reset_index()
            super_summary.rename(columns=io.to_camel).to_excel(
                writer, sheet_name="Superfamily_summary", index=False)

        if dl is not None:
            loc_df = df[["protein_id", "family"]].drop_duplicates().merge(
                dl, on="protein_id", how="left"
            )
            loc_df.rename(columns=io.to_camel).to_excel(writer, sheet_name="Localization", index=False)

    external.log("[OK] Final XLSX saved: gwiscan_results.xlsx")
    external.log("=" * 60)
    external.log("GWIscan complete.")
    external.log(f"  Proteins identified : {df['protein_id'].nunique()}")
    external.log(f"  Domain hits total   : {len(df)}")
    external.log(f"  Families detected   : {df['family'].nunique()}")
    external.log(f"  Output directory    : {cfg.final_dir}")
    external.log("=" * 60)
