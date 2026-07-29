"""Tests for compile: Pfam domain-architecture string + standalone/multidomain flag.

Architecture uses Pfam only (one coordinate system), ordered N->C by start.
"""

import pandas as pd

from gwiscan.compile import architecture_type, pfam_architecture


def _ipr():
    return pd.DataFrame([
        # P1: out-of-order rows + a CDD row that must be ignored for architecture.
        {"protein_id": "P1", "analysis": "Pfam", "sig_acc": "PF00069.27", "sig_desc": "Pkinase", "start": 300},
        {"protein_id": "P1", "analysis": "Pfam", "sig_acc": "PF01453.31", "sig_desc": "B_lectin", "start": 40},
        {"protein_id": "P1", "analysis": "Pfam", "sig_acc": "PF00024",    "sig_desc": "PAN_1",    "start": 200},
        {"protein_id": "P1", "analysis": "CDD",  "sig_acc": "cd00192",    "sig_desc": "STKc",     "start": 305},
        # P2: two copies of the family domain only.
        {"protein_id": "P2", "analysis": "Pfam", "sig_acc": "PF01453",    "sig_desc": "B_lectin", "start": 10},
        {"protein_id": "P2", "analysis": "Pfam", "sig_acc": "PF01453",    "sig_desc": "B_lectin", "start": 150},
    ])


def test_architecture_ordered_n_to_c_pfam_only():
    arch = pfam_architecture(_ipr())
    # sorted by start; the CDD row is excluded.
    assert [label for _, label in arch["P1"]] == ["B_lectin", "PAN_1", "Pkinase"]
    assert [acc for acc, _ in arch["P1"]] == ["PF01453", "PF00024", "PF00069"]


def test_type_standalone_vs_multidomain():
    arch = pfam_architecture(_ipr())
    p1 = [acc for acc, _ in arch["P1"]]
    p2 = [acc for acc, _ in arch["P2"]]
    # P1 (family GNA = PF01453) also carries kinase + PAN -> multidomain, 1 family copy.
    assert architecture_type(p1, "PF01453") == (1, "multidomain")
    # P2 has only two B_lectin copies -> standalone, count 2.
    assert architecture_type(p2, "PF01453") == (2, "standalone")
    # no Pfam evidence at all.
    assert architecture_type([], "PF01453") == (0, "-")


def test_pfam_architecture_handles_empty():
    assert pfam_architecture(None) == {}
    assert pfam_architecture(pd.DataFrame()) == {}


# --- XLSX grouped (two-row) header: band_spans + _apply_band_header -----------

from gwiscan import compile as compile_mod
from gwiscan.compile import band_spans


def _label(col):
    """The band label for a column -- derived from COLUMN_BANDS so these tests
    track label-text edits instead of breaking on them."""
    return next(lbl for lbl, cols in compile_mod.COLUMN_BANDS if col in cols)


def test_band_spans_groups_contiguous_columns():
    cols = ["protein_id", "family", "gene_id", "chrom", "accession", "evalue"]
    spans = band_spans(cols)
    assert spans == [
        (_label("protein_id"), 0, 1),
        (_label("gene_id"), 2, 3),
        (_label("accession"), 4, 5),
    ]


def test_band_spans_tiles_full_width_exactly_once():
    cols = ["protein_id", "family", "gene_id", "chrom", "gene_start", "gene_end",
            "strand", "accession", "evalue", "bitscore", "start", "end", "method"]
    spans = band_spans(cols)
    covered = [i for _, a, b in spans for i in range(a, b + 1)]
    assert covered == list(range(len(cols)))   # no gaps, no overlaps, in order


def test_band_spans_unknown_column_gets_blank_own_span():
    cols = ["protein_id", "mystery_col", "family"]
    spans = band_spans(cols)
    # mystery_col isn't in any band: its own blank span, family starts a new one
    assert spans == [("Identity", 0, 0), ("", 1, 1), ("Identity", 2, 2)]


def test_band_spans_irregular_targetp_columns_match():
    # noTP/SP/mTP/cTP/luTP are preserved (not snake_case); they must still band.
    cols = ["targetp_type", "noTP", "SP", "mTP", "cTP", "luTP", "cs_position"]
    spans = band_spans(cols)
    assert spans == [(_label("noTP"), 0, 6)]   # one band spanning all TargetP cols


def test_all_real_columns_are_mapped():
    # every column the pipeline can emit belongs to some band (no silent blanks)
    labeled = {c for _, cols in compile_mod.COLUMN_BANDS for c in cols}
    real = ["protein_id", "family", "superfamily", "gene_id", "chrom", "gene_start",
            "gene_end", "strand", "accession", "evalue", "bitscore", "start", "end",
            "method", "domain_architecture", "family_domain_count", "architecture_type",
            "evidence_level", "evidence_support", "evidence_criteria", "length_aa",
            "molecular_weight", "isoelectric_point", "negatively_charged_residues",
            "positively_charged_residues", "instability_index", "gravy", "aliphatic_index",
            "ec_cystines", "ec_reduced", "targetp_type", "noTP", "SP", "mTP", "cTP",
            "luTP", "cs_position", "topology", "signal_peptide", "tm_regions",
            "n_tm_regions", "beta_regions", "localizations", "signals", "membrane_types",
            "ipr_accessions", "ipr_descriptions", "go_terms", "go_term_names",
            "go_molecular_function", "go_biological_process", "go_cellular_component",
            "analyses_hit"]
    assert set(real) <= labeled


def test_apply_band_header_writes_two_rows_and_freezes(tmp_path):
    import pandas as pd
    from openpyxl import load_workbook
    from gwiscan import io

    cols = ["protein_id", "family", "accession", "evalue", "length_aa", "molecular_weight"]
    df = pd.DataFrame([["a", "b", "c", 1.0, 100, 12.3]], columns=cols)
    out = tmp_path / "r.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        spans = band_spans(df.columns)
        df.rename(columns=io.to_camel).to_excel(w, sheet_name="All_candidates", index=False, startrow=1)
        compile_mod._apply_band_header(w.sheets["All_candidates"], spans)

    ws = load_workbook(out)["All_candidates"]
    assert ws["A1"].value == _label("protein_id")       # band row
    assert ws["C1"].value == _label("accession")
    assert ws["E1"].value == _label("length_aa")
    assert ws["A2"].value == "proteinId"                # column-name row (camel)
    assert ws["C2"].value == "accession"
    assert ws["A3"].value == "a"                         # data starts row 3
    assert ws.freeze_panes == "C3"
    assert len(ws.merged_cells.ranges) == 3             # one merge per multi-col band
